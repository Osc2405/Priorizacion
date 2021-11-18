import pandas as pd
import numpy as np
import openpyxl
import os


path="EquiposHUV-Zona4-copia.xlsx"
df=pd.read_excel(path)

df_salida=pd.DataFrame()
columnas_originales=df[["Nombre","Descripcion adicional","Marca","Modelo","Estado actual","Servicio","Clasificación biomedica","Clasificación de riesgo", "Cantidad de correctivos registrados"]]
df_salida=columnas_originales.copy()
df_salida["Estado ponderado"]=np.where(df_salida["Estado actual"]=="Activo",1,np.where(df_salida["Estado actual"]=="Fuera de servicio",1,0))
path2="datos/servicios.csv"
df_servicios=pd.read_csv(path2,sep=";",encoding='latin-1')
servicios_dic = df_servicios.set_index('Servicio').to_dict()["Peso"]
df_salida["Servicio ponderado"]=df_salida["Servicio"].apply(lambda x: servicios_dic[x])

clasificacion_dic={"TRATAMIENTO Y MANTENIMIENTO DE LA VIDA":3,
                    "RAHABILITACIÓN":2,"DIAGNOSTICO":2,
                    "APOYO":1,"N/R":1,"ANÁLISIS DE LABORATORIO":1,"REHABILITACIÓN":1,"NaN":1}
riesgo_dic={"I":1,"IIA":2,"IIB":3,"III":4,"NaN":1,"N/R":1}

df_salida["Clasificación biomedica"]=df_salida["Clasificación biomedica"].fillna("NaN")
df_salida["Clasificación de riesgo"]=df_salida["Clasificación de riesgo"].fillna("NaN")
df_salida["Clasificacion ponderada"]=df_salida["Clasificación biomedica"].apply(lambda x: clasificacion_dic[x])
df_salida["Riesgo ponderado"]=df_salida["Clasificación de riesgo"].apply(lambda x: riesgo_dic[x])

def func(x):
    if int(x)<=5:
        return 1
    elif int(x)<=9:
        return 2
    elif int(x)>10:
        return 3
    else:
        return 1

df_salida["Correctivos ponderados"]=df_salida["Cantidad de correctivos registrados"].apply(func)
df_salida["Impacto operacional"]=1
df_salida["Solicitud por el personal"]=1


# def total(x):
#     salida=x[9]*(x[10]*0.211+x[11]*0.033+x[12]*0.128+x[13]*0.06+x[14]*0.208+x[15]*0.359)
#     return salida

# df_salida["Total"]=df_salida.apply(total,axis=1)
df_salida.to_excel("Datos_filtrados.xlsx")

wbk=openpyxl.load_workbook("Datos_filtrados.xlsx")
for wks in wbk.worksheets:
    wks.cell(row=1,column=18).value="Total"
wbk.save("Datos_filtrados.xlsx")
wbk.close()
wbk=openpyxl.load_workbook("Datos_filtrados.xlsx")
wks=wbk.worksheets[0]
maxRow=wks.max_row
for row in range(2,maxRow):
    wks.cell(row=row,column=18).value="=K{}*(L{}*0.211+M{}*0.033+N{}*0.128+O{}*0.06+P{}*0.208+Q{}*0.359)".format(row,row,row,row,row,row,row)
wbk.save("Datos_filtrados.xlsx")
wbk.close()

