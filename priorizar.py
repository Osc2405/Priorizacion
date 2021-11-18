from tkinter import Tk, Label, Button, Frame,  messagebox, filedialog, ttk, Scrollbar, VERTICAL, HORIZONTAL
import pandas as pd
import numpy as np
import openpyxl
import os


ventana = Tk()
ventana.config(bg='black')
ventana.geometry('600x400')
ventana.minsize(width=600, height=400)
ventana.title('Leer datos de Excel')

ventana.columnconfigure(0, weight = 25)
ventana.rowconfigure(0, weight= 25)
ventana.columnconfigure(0, weight = 1)
ventana.rowconfigure(1, weight= 1)

frame1 = Frame(ventana, bg='gray26')
frame1.grid(column=0,row=0,sticky='nsew')
frame2 = Frame(ventana, bg='gray26')
frame2.grid(column=0,row=1,sticky='nsew')

frame1.columnconfigure(0, weight = 1)
frame1.rowconfigure(0, weight= 1)

frame2.columnconfigure(0, weight = 1)
frame2.rowconfigure(0, weight= 1)
frame2.columnconfigure(1, weight = 1)
frame2.rowconfigure(0, weight= 1)

frame2.columnconfigure(2, weight = 1)
frame2.rowconfigure(0, weight= 1)

frame2.columnconfigure(3, weight = 2)
frame2.rowconfigure(0, weight= 1)

frame2.columnconfigure(4, weight = 2)
frame2.rowconfigure(0, weight= 1)

def func(x):
    if int(x)<=5:
        return 1
    elif int(x)<=9:
        return 2
    elif int(x)>10:
        return 3
    else:
        return 1

def abrir_archivo():

	archivo = filedialog.askopenfilename(initialdir ='/', 
											title='Selecione archivo', 
											filetype=(('xlsx files', '*.xlsx*'),('All files', '*.*')))
	indica['text'] = archivo


def datos_excel():

	datos_obtenidos = indica['text']
	try:
		archivoexcel = r'{}'.format(datos_obtenidos)
		

		df = pd.read_excel(archivoexcel)

	except ValueError:
		messagebox.showerror('Informacion', 'Formato incorrecto')
		return None

	except FileNotFoundError:
		messagebox.showerror('Informacion', 'El archivo esta \n malogrado')
		return None

	Limpiar()

	tabla['column'] = list(df.columns)
	tabla['show'] = "headings"  #encabezado
     

	for columna in tabla['column']:
		tabla.heading(columna, text= columna)
	

	df_fila = df.to_numpy().tolist()
	for fila in df_fila:
		tabla.insert('', 'end', values =fila)


def Limpiar():
	tabla.delete(*tabla.get_children())

def Procesar():
    print("Hola")
    datos_obtenidos = indica['text']
    try:
        archivoexcel = r'{}'.format(datos_obtenidos)
        df = pd.read_excel(archivoexcel)
    
    except ValueError:
        messagebox.showerror('Informacion', 'Formato incorrecto')
        return None
    
    except FileNotFoundError:
        messagebox.showerror('Informacion', 'El archivo esta \n malogrado')
        return None

    df_salida=pd.DataFrame()
    columnas_originales=df[["Nombre","Descripcion adicional","Marca","Modelo","Estado actual","Servicio","Clasificación biomedica","Clasificación de riesgo", "Cantidad de correctivos registrados"]]
    df_salida=columnas_originales.copy()
    df_salida["Estado ponderado"]=np.where(df_salida["Estado actual"]=="Activo",1,np.where(df_salida["Estado actual"]=="Fuera de servicio",1,0))
    path2="datos/servicios.csv"
    df_servicios=pd.read_csv(path2,sep=";",encoding='latin-1')
    servicios_dic = df_servicios.set_index('Servicio').to_dict()["Peso"]
    df_salida["Servicio ponderado"]=df_salida["Servicio"].apply(lambda x: servicios_dic[x])

    clasificacion_dic={"TRATAMIENTO Y MANTENIMIENTO DE LA VIDA":3,
                    "RAHABILITACIÓN":2,"DIAGNOSTICO":2,
                    "APOYO":1,"N/R":1,"ANÁLISIS DE LABORATORIO":1,"REHABILITACIÓN":1,"NaN":1}
    riesgo_dic={"I":1,"IIA":2,"IIB":3,"III":4,"NaN":1,"N/R":1}

    df_salida["Clasificación biomedica"]=df_salida["Clasificación biomedica"].fillna("NaN")
    df_salida["Clasificación de riesgo"]=df_salida["Clasificación de riesgo"].fillna("NaN")
    df_salida["Clasificacion ponderada"]=df_salida["Clasificación biomedica"].apply(lambda x: clasificacion_dic[x])
    df_salida["Riesgo ponderado"]=df_salida["Clasificación de riesgo"].apply(lambda x: riesgo_dic[x])

    

    df_salida["Correctivos ponderados"]=df_salida["Cantidad de correctivos registrados"].apply(func)
    df_salida["Impacto operacional"]=1
    df_salida["Solicitud por el personal"]=1


    # def total(x):
    #     salida=x[9]*(x[10]*0.211+x[11]*0.033+x[12]*0.128+x[13]*0.06+x[14]*0.208+x[15]*0.359)
    #     return salida

    # df_salida["Total"]=df_salida.apply(total,axis=1)
    df_salida.to_excel("Datos_filtrados.xlsx")

    wbk=openpyxl.load_workbook("Datos_filtrados.xlsx")
    for wks in wbk.worksheets:
        wks.cell(row=1,column=18).value="Total"
    wbk.save("Datos_filtrados.xlsx")
    wbk.close()
    wbk=openpyxl.load_workbook("Datos_filtrados.xlsx")
    wks=wbk.worksheets[0]
    maxRow=wks.max_row
    for row in range(2,maxRow):
        wks.cell(row=row,column=18).value="=K{}*(L{}*0.211+M{}*0.033+N{}*0.128+O{}*0.06+P{}*0.208+Q{}*0.359)".format(row,row,row,row,row,row,row)
    wbk.save("Datos_filtrados.xlsx")
    wbk.close()


    print ("Hola")


tabla = ttk.Treeview(frame1 , height=10)
tabla.grid(column=0, row=0, sticky='nsew')

ladox = Scrollbar(frame1, orient = HORIZONTAL, command= tabla.xview)
ladox.grid(column=0, row = 1, sticky='ew') 

ladoy = Scrollbar(frame1, orient =VERTICAL, command = tabla.yview)
ladoy.grid(column = 1, row = 0, sticky='ns')

tabla.configure(xscrollcommand = ladox.set, yscrollcommand = ladoy.set)

estilo = ttk.Style(frame1)
estilo.theme_use('clam') #  ('clam', 'alt', 'default', 'classic')
estilo.configure(".",font= ('Arial', 14), foreground='red2')
estilo.configure("Treeview", font= ('Helvetica', 12), foreground='black',  background='white')
estilo.map('Treeview',background=[('selected', 'green2')], foreground=[('selected','black')] )


boton1 = Button(frame2, text= 'Abrir', bg='lightBlue1', command= abrir_archivo)
boton1.grid(column = 0, row = 0, sticky='nsew', padx=10, pady=10)

boton2 = Button(frame2, text= 'Mostrar', bg='lightBlue2', command= datos_excel)
boton2.grid(column = 1, row = 0, sticky='nsew', padx=10, pady=10)

boton3 = Button(frame2, text= 'Limpiar', bg='lightBlue3', command= Limpiar)
boton3.grid(column = 2, row = 0, sticky='nsew', padx=10, pady=10)

boton3 = Button(frame2, text= 'Procesar', bg='lightBlue4', command= Procesar)
boton3.grid(column = 3, row = 0, sticky='nsew', padx=10, pady=10)


indica = Label(frame2, fg= 'white', bg='gray26', text= 'Ubicación del archivo', font= ('Arial',10,'bold') )
indica.grid(column=4, row = 0)

ventana.mainloop()

